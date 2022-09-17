import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from objects import Household, Flat
from objects.Allocation import Allocation
from objects.HappyNumbers import HappyNumbers


def read_source(file, file2, list_hh_wishes, list_flats, list_weights, weights_from_first=False):

    wb = load_workbook(file, data_only=True)
    hh_wishes = wb.worksheets[0]
    last_row = int(re.search(r'\d*$', hh_wishes.dimensions).group(0))
    for row in range(2, last_row + 1):
        list_hh_wishes[hh_wishes["A" + str(row)].value] = \
            Household.Household(
                hh_wishes["A" + str(row)].value,  # id
                hh_wishes["G" + str(row)].value,  # name
                hh_wishes["J" + str(row)].value,  # flat-type
                hh_wishes["BA" + str(row)].value,  # wbs
                hh_wishes["BB" + str(row)].value,  # engagement

                hh_wishes["K" + str(row)].value,  # punkt, ...
                hh_wishes["L" + str(row)].value,
                hh_wishes["M" + str(row)].value,
                hh_wishes["N" + str(row)].value,
                hh_wishes["O" + str(row)].value,
                hh_wishes["P" + str(row)].value,

                hh_wishes["Q" + str(row)].value,  # eg, ...
                hh_wishes["R" + str(row)].value,
                hh_wishes["S" + str(row)].value,
                hh_wishes["T" + str(row)].value,
                hh_wishes["U" + str(row)].value,
                hh_wishes["V" + str(row)].value,
                hh_wishes["W" + str(row)].value,
                hh_wishes["X" + str(row)].value,

                hh_wishes["Y" + str(row)].value,  # small_flat

                hh_wishes["Z" + str(row)].value,  # wheelchair
                hh_wishes["AA" + str(row)].value,

                hh_wishes["AB" + str(row)].value,  # neighbour
                hh_wishes["AD" + str(row)].value,
                hh_wishes["AE" + str(row)].value,

                hh_wishes["AF" + str(row)].value,  # dog
                hh_wishes["AG" + str(row)].value,
                hh_wishes["AH" + str(row)].value,
                hh_wishes["AI" + str(row)].value,
                hh_wishes["AJ" + str(row)].value,  # cat
                hh_wishes["AK" + str(row)].value,
                hh_wishes["AL" + str(row)].value,
                hh_wishes["AM" + str(row)].value,

                hh_wishes["AN" + str(row)].value,  # smoker
                hh_wishes["AO" + str(row)].value,

                hh_wishes["AP" + str(row)].value,  # flat_wish
                hh_wishes["AQ" + str(row)].value,
                hh_wishes["AR" + str(row)].value,
                hh_wishes["AS" + str(row)].value
            )

        if list_hh_wishes[hh_wishes["A" + str(row)].value].engagement is None:
            list_hh_wishes[hh_wishes["A" + str(row)].value].engagement = 0

    wb2 = load_workbook(file2, data_only=True)
    ws_flat_data = wb2.worksheets[0]
    last_row_wohndaten = int(re.search(r'\d*$', ws_flat_data.dimensions).group(0))
    for row in range(2, last_row_wohndaten + 1):
        list_flats[ws_flat_data["A" + str(row)].value] = \
            Flat.Flat(
                ws_flat_data["A" + str(row)].value,  # WohnungsID
                ws_flat_data["B" + str(row)].value,  # Art
                ws_flat_data["C" + str(row)].value,  # Gebäude
                ws_flat_data["D" + str(row)].value,  # Etage
                ws_flat_data["E" + str(row)].value,  # qm
                ws_flat_data["F" + str(row)].value,  # Kleine_Wg
                ws_flat_data["G" + str(row)].value,  # WBS
                ws_flat_data["H" + str(row)].value,  # Rollitauglich
                ws_flat_data["I" + str(row)].value  # Vergabe
            )

    if not weights_from_first:
        config = wb2.worksheets[1]
        first_row_gew = int(re.search(r'[0-9]{1,4}(?=:)', config.tables["gewichte_tab"].ref).group()) + 1
        first_col_gew = re.search(r'^[A-Z]{1,4}', config.tables["gewichte_tab"].ref).group()
        last_row_gew = int(re.search(r'[0-9]{1,4}$', config.tables["gewichte_tab"].ref).group())
        last_col_gew = re.search(r'(?<=:)[A-Z]*', config.tables["gewichte_tab"].ref).group()
        for row in range(first_row_gew, last_row_gew + 1):
            list_weights[config[first_col_gew + str(row)].value] = float(config[last_col_gew + str(row)].value)
    else:
        ws_alloc = wb["Zuordnungen"]
        for col in range(5,25):
            list_weights[ws_alloc[get_column_letter(col) + str(1)].value] = float(ws_alloc[get_column_letter(col) + str(2)].value)

def read_results(file, allocations):
    wb = load_workbook(file)

    ws_alloc = wb["Zuordnungen"]
    last_row = int(re.search(r'\d*$', ws_alloc.dimensions).group(0))
    for row in range(5, last_row + 1):
        allocations[ws_alloc["A" + str(row)].value] = \
            Allocation(
                ws_alloc["A" + str(row)].value,
                ws_alloc["B" + str(row)].value
            )
        if int(ws_alloc["A" + str(row)].value) < 1000:
            allocations[ws_alloc["A" + str(row)].value].happy_numbers = \
                HappyNumbers(ws_alloc["C" + str(row)].value,
                             ws_alloc["D" + str(row)].value,
                             ws_alloc["E" + str(row)].value,
                             ws_alloc["F" + str(row)].value,
                             ws_alloc["G" + str(row)].value,
                             ws_alloc["H" + str(row)].value,
                             ws_alloc["I" + str(row)].value)
        else:
            allocations[ws_alloc["A" + str(row)].value].happy_numbers = \
                HappyNumbers(0, 0, 0, 0, 0, 0, 0)

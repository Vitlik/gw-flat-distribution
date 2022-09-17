import pickle
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def save_data_to_xlsx(file, hh_wishes, flats, allocations, weights, max_happiness, swapping):
    wb = load_workbook(file)

    sheet_name = "Zuordnungen" + swapping
    ws = wb.create_sheet(sheet_name)

    # wechsle aktiven Tab
    wb.active = wb[sheet_name]
    for sheet in wb:
        if sheet.title == sheet_name:
            sheet.sheet_view.tabSelected = True
        else:
            sheet.sheet_view.tabSelected = False

    ws.append(["Maximales Glück:", "", "", "Gewichte:", "Punkt",	"Winkel",	"Riegel",	"EG",	"1OG",	"2OG",	"3OG",
              "Nachbar",	"kleine_wg",	"Hund",	"Hundeallergie",	"Katze",	"Katzenallergie",	"konkrete_wg1",
              "konkrete_wg2",	"konkrete_wg3",	"rollitauglich",	"Raucher",	"WBS-AB",	"WBS-AB"])
    ws.append([round(max_happiness, 4), "", "", "", weights["Punkt"],	weights["Winkel"],	weights["Riegel"],
              weights["EG"],	weights["1OG"],	weights["2OG"],	weights["3OG"],	weights["Nachbar"],	weights["kleine_wg"],
              weights["Hund"],	weights["Hundeallergie"],	weights["Katze"],	weights["Katzenallergie"],
              weights["konkrete_wg1"],	weights["konkrete_wg2"],	weights["konkrete_wg3"],	weights["rollitauglich"],
              weights["Raucher"],	weights["WBS-AB"],	weights["EngagementMultiplikator"]])
    ws.append([""])
    tabHead = Table(displayName="weights", ref="D1:" + get_column_letter(23) + "2")

    # Lege Tabellen Stil fest (striped rows and banded columns)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tabHead.tableStyleInfo = style

    ws["A2"].alignment = Alignment(horizontal='center')
    ws["A2"].fill = PatternFill("solid", start_color="92D050")

    ws.append(["HaushaltsID", "WohnungsID", "Name", "Größe", "Summe",
               "Punkt_Glück", "Winkel_Glück", "Riegel_Glück", "EG_Glück", "OG1_Glück", "OG2_Glück", "OG3_Glück",
               "kleiWg_Glück", "Rolli_Glück", "Nachbar_Glück", "HundNähe_Glück", "KatzenNähe_Glück", "RaucherNähe_Glück",
               "Wg_Glück", "WBS", " ",
               "1. unerfüllter wichtiger Wunsch (4 oder 5)", "2. unerfüllter wichtiger Wunsch (4 oder 5)",
               "3. unerfüllter wichtiger Wunsch (4 oder 5)", "4. unerfüllter wichtiger Wunsch (4 oder 5)",
               "5. unerfüllter wichtiger Wunsch (4 oder 5)",
               "WBS Änderung"])

    counter = 4
    for alloc in allocations:
        if alloc < 1000:
            ws.append(
                [allocations[alloc].hh_id,
                 allocations[alloc].wg_id,
                 hh_wishes[alloc].name,
                 flats[allocations[alloc].wg_id].flat_type,
                 # sum column
                 allocations[alloc].happy_numbers.sum,
                 # building wish
                 allocations[alloc].happy_numbers.punkt,
                 allocations[alloc].happy_numbers.winkel,
                 allocations[alloc].happy_numbers.riegel,
                 # floor wish
                 allocations[alloc].happy_numbers.eg,
                 allocations[alloc].happy_numbers.og1,
                 allocations[alloc].happy_numbers.og2,
                 allocations[alloc].happy_numbers.og3,
                 # small flat wisch
                 allocations[alloc].happy_numbers.small_flat,
                 # wheelchair wisch
                 allocations[alloc].happy_numbers.wheelchair,
                 # neighbour wish
                 allocations[alloc].happy_numbers.neighbour,
                 # animal wish
                 allocations[alloc].happy_numbers.dog,
                 allocations[alloc].happy_numbers.cat,
                 # smoker wisch
                 allocations[alloc].happy_numbers.smoker,
                 # specific flat wisch
                 allocations[alloc].happy_numbers.specific_flat,
                 # fitting wbs allocation
                 allocations[alloc].happy_numbers.wbs,
                 "",
                 # unfulfilled wishes
                 allocations[alloc].happy_numbers.unfulfilled_wish1,
                 allocations[alloc].happy_numbers.unfulfilled_wish2,
                 allocations[alloc].happy_numbers.unfulfilled_wish3,
                 allocations[alloc].happy_numbers.unfulfilled_wish4,
                 allocations[alloc].happy_numbers.unfulfilled_wish5,
                 # WBS Change
                 allocations[alloc].happy_numbers.wbs_change
                 ])
        else:
            ws.append([allocations[alloc].hh_id, allocations[alloc].wg_id])
        counter += 1

    # Formatiere Daten als Tabelle
    if not swapping:
        tab = Table(displayName="Allocations", ref="A4:" + get_column_letter(27) + str(counter))
    else:
        tab = Table(displayName="Allocations_neu", ref="A4:" + get_column_letter(27) + str(counter))

    # Lege Tabellen Stil fest (striped rows and banded columns)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    # füge neues Datenblatt zur Datei hinzu
    ws.add_table(tab)

    # lege Spaltenbreite pauschal fest
    for i in range(1, 30):  # ,1 to start at 1
        ws.column_dimensions[get_column_letter(i)].width = 10

    # Färbe Spalte grau
    for cell in ws['U5:U' + str(counter)]:
        cell[0].fill = PatternFill("solid", start_color="D9D9D9")
    # ws.column_dimensions["T"].width = 3

    # lege Dateinamen fest
    if not swapping:
        new_file_name = file.replace(".xlsx", "") + "_" + datetime.now().strftime("%Y-%m-%d_%H-%M-%S") \
                        + "_" + str(round(max_happiness, 4)) + ".xlsx"
    else:
        new_file_name = file.replace(".xlsx", "") + "_" + swapping + ".xlsx"

    wb.save(new_file_name)
    print("File saved: " + new_file_name)


def save_allocation(allocations, suffix):
    with open("datasources/allocations_" + datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + "_" + suffix + ".pkl", "wb") as out:
        pickle.dump(allocations, out, pickle.HIGHEST_PROTOCOL)
